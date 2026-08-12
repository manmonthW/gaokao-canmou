#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""双万计划（一流本科专业建设"双万计划"）采集 → etl/data/swyc_batches.csv

数据来源与背景
==============
教育部三批名单（教高厅函〔2019〕46号/〔2021〕7号/〔2022〕14号）采用"分送"方式下发：
各省只收到本省名单、部属高校只收到本校名单，官方从未公布全国统一名单，
MOE 官网附件亦需验证码下载。因此本采集采用两路数据：

1) 国家级全量（主体，11k+ 条）：
   中国教育在线·掌上高考 static-data 接口
   https://static-data.eol.cn/www/2.0/school/{school_id}/pc_special.json
   字段 nation_first_class="1" 即"国家级一流本科专业建设点"（三批累计）。
   原始 JSON 已存档 major/swyc/eol_pc_special/<school_id>.json（2991 校，
   抓取脚本见 git 历史/etl 暂存，2026-08-12 采集，限速 0.15s+8 并发）。
   学校名映射：major/swyc/eol_name.json。
   该源无法区分入选批次（batch/batch_year 置空）。

2) 批次标注增强（广东省 2021 批次，官方附件 xlsx）：
   东莞理工教务部转载的教高厅函〔2022〕14号广东省附件
   https://jwb.dgut.edu.cn/info/1271/31371.htm
   存档 major/swyc/dgut_2021_guangdong_lists.xlsx（sheet 国家级/省级）。
   用于给匹配行补 batch=3, batch_year=2021，并新增省级(provincial)行。

列：school_name, major_name, batch(1/2/3 或空), batch_year(2019/2020/2021 或空), tier(national/provincial)

用法: python3 etl/collect_swyc.py [--offline]
"""
import csv, json, os, sys
from collections import Counter

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
SPDIR = os.path.join(ROOT, "major", "swyc", "eol_pc_special")
NAME_JSON = os.path.join(ROOT, "major", "swyc", "eol_name.json")
GD_XLSX = os.path.join(ROOT, "major", "swyc", "dgut_2021_guangdong_lists.xlsx")
OUT = os.path.join(BASE, "data", "swyc_batches.csv")


def load_eol_national():
    """从归档 JSON 提取国家级一流本科专业（三批累计）。"""
    names = {s["school_id"]: s["name"]
             for s in json.load(open(NAME_JSON, encoding="utf-8"))["data"]}
    rows = {}  # (school, major) -> code
    for fn in sorted(os.listdir(SPDIR)):
        sid = fn[:-5]
        try:
            data = json.load(open(os.path.join(SPDIR, fn), encoding="utf-8")).get("data") or {}
        except Exception:
            continue
        majors = []
        for grp in (data.get("special_detail") or {}).values():
            if isinstance(grp, list):
                majors.extend(m for m in grp if isinstance(m, dict))
        if not majors:
            for k in ("1", "2"):
                v = data.get(k)
                if isinstance(v, list):
                    majors.extend(m for m in v if isinstance(m, dict))
        for m in majors:
            if m.get("nation_first_class") != "1":
                continue
            if not str(m.get("type_name", "")).startswith("本科"):
                continue
            sname = names.get(sid)
            mname = (m.get("special_name") or "").strip()
            if not sname or not mname:
                continue
            rows.setdefault((sname, mname), m.get("code", ""))
    return rows


def load_guangdong_2021():
    """解析 dgut 归档的教高厅函〔2022〕14号广东省附件 xlsx。"""
    import openpyxl
    wb = openpyxl.load_workbook(GD_XLSX, data_only=True)
    nat, prov = set(), set()
    for sheet, target in (("国家级", nat), ("省级", prov)):
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(values_only=True):
            if not row or len(row) < 3:
                continue
            seq, school, major = row[0], row[1], row[2]
            if not isinstance(seq, (int, float)) or not school or not major:
                continue
            target.add((str(school).strip(), str(major).strip()))
    return nat, prov


def main():
    nat_rows = load_eol_national()
    print("eol 国家级（三批累计）:", len(nat_rows))
    gd_nat, gd_prov = load_guangdong_2021()
    print("广东2021 国家级:", len(gd_nat), "省级:", len(gd_prov))

    out = []
    matched = 0
    for (school, major), _code in nat_rows.items():
        batch, year = "", ""
        if (school, major) in gd_nat:
            batch, year, matched = "3", "2021", matched + 1
        out.append((school, major, batch, year, "national"))
    # 广东 2021 国家级中 eol 未覆盖的（少量差异）补入
    eol_keys = set(nat_rows)
    extra = gd_nat - eol_keys
    for school, major in sorted(extra):
        out.append((school, major, "3", "2021", "national"))
    # 广东 2021 省级
    for school, major in sorted(gd_prov):
        out.append((school, major, "3", "2021", "provincial"))
    print("广东2021国家级与eol匹配:", matched, " eol未覆盖补入:", len(extra))

    out.sort(key=lambda r: (r[4], r[0], r[1]))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["school_name", "major_name", "batch", "batch_year", "tier"])
        w.writerows(out)
    tc = Counter(r[4] for r in out)
    bc = Counter(r[2] for r in out)
    print("写入:", OUT, "rows:", len(out), "tier:", dict(tc), "batch:", dict(bc))
    schools = len({r[0] for r in out})
    print("涉及学校:", schools)


if __name__ == "__main__":
    main()

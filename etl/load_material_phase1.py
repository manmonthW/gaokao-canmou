#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
load_material_phase1.py  —— 第一期：院校+专业核心字段补全（本地资料库）

覆盖：
  A1  全国普通高等学校名单.xls(2026) -> school_profiles(school_id_code/affiliation/province/city/level)
  A2  985/211/双一流/小985/小211/部委直属 PDF -> school_profiles(is_985/is_211/is_dfc/tags)
  A3  全国具有保研资格的高校367所.pdf -> school_profiles(has_postgrad_recommend)
  B6  2027lnzsxkap0407jx.xlsx(辽宁选考科目全量) -> major_profiles(subject_req)
  B7  普通高等学校本科专业目录2026.pdf -> major_catalog(专业字典)

所有匹配均以"院校名称"为主键（schools.code 为4位辽宁代码，与材料代码格式不同，用名称对齐）。
用法：
  python3 load_material_phase1.py --all        # 全部执行
  python3 load_material_phase1.py --a1
  python3 load_material_phase1.py --a2
  python3 load_material_phase1.py --a3
  python3 load_material_phase1.py --b6
  python3 load_material_phase1.py --b7
  python3 load_material_phase1.py --report     # 写入后打印覆盖率
"""
import argparse, os, re, glob, sys
import pdfplumber
import openpyxl, xlrd
import psycopg2

BASE = "/home/ekewang/projects/gaokao/ln/2026allmaterial"
DB = dict(host="localhost", port=5432, dbname="gaokao", user="gaokao", password="gaokao123")


def conn_db():
    return psycopg2.connect(**DB)


# ---------- A1: 普通高校名单.xls ----------
def load_a1(conn):
    path = os.path.join(BASE, "全国普通高等学校名单（截止2026年6月17日）.xls")
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        vals = sh.row_values(r)
        name = str(vals[1]).strip() if vals[1] else ""
        if not name or "附件" in name or "全国普通高等学校名单" in name:
            continue
        if name in ("序号", "学校名称") or re.match(r"^[\u4e00-\u9fa5]+（\d+所）$", name):
            continue  # 表头/省市分隔行
        # 学校标识码：xlrd 读成浮点，转纯数字串
        cid_raw = vals[2]
        code_id = None
        if cid_raw:
            try:
                code_id = str(int(float(cid_raw)))
            except Exception:
                code_id = str(cid_raw).strip()
        dept = str(vals[3]).strip() if vals[3] else None
        loc = str(vals[4]).strip() if vals[4] else None
        level = str(vals[5]).strip() if vals[5] else None
        rows.append((name, code_id, dept, loc, level))
    print(f"[A1] 读取名单 {len(rows)} 行")

    cur = conn.cursor()
    # 建立 名称->code 映射（schools）
    cur.execute("SELECT code, name FROM schools")
    name2code = {n: c for c, n in cur.fetchall()}
    updated = 0
    for name, code_id, dept, loc, level in rows:
        code = name2code.get(name)
        if not code:
            continue
        # 所在地拆 省/市：格式如 "辽宁省沈阳市" / "北京市"（直辖市省=市）
        province = loc
        city = None
        if loc:
            m = re.match(r"^(.+?(?:省|自治区|市))(.*)$", loc)
            if m:
                province = m.group(1)
                rest = m.group(2).strip()
                city = rest if rest else province  # 直辖市 city=province
        # city 去掉"市"后缀匹配 cities 表；匹配不上则 NULL（避免 FK 违规）
        city_key = city.replace("市", "") if city else None
        if city_key:
            cur.execute("SELECT 1 FROM cities WHERE city=%s", (city_key,))
            if not cur.fetchone():
                city_key = None
        cur.execute("""
            UPDATE school_profiles
            SET school_id_code=%s, affiliation=%s, province=%s, city=%s, level=%s, enriched_at=now()
            WHERE code=%s
        """, (code_id, dept, province, city_key, level, code))
        updated += 1
    conn.commit()
    print(f"[A1] 更新 school_profiles {updated} 条（基于名称匹配）")


# ---------- A2: 层级名单 PDF ----------
def _pdf_lines(pdf_path, max_pages=10):
    out = []
    with pdfplumber.open(pdf_path) as pdf:
        for pi in range(min(max_pages, len(pdf.pages))):
            t = pdf.pages[pi].extract_text() or ""
            out.extend(t.splitlines())
    return out


def load_a2(conn):
    cur = conn.cursor()
    cur.execute("SELECT code, name FROM schools")
    name2code = {n: c for c, n in cur.fetchall()}
    sub = os.path.join(BASE, "2025高考志愿填报资料汇总")

    def names_in_pdf(fname, label):
        lines = _pdf_lines(os.path.join(sub, fname), max_pages=20)
        found = set()
        for ln in lines:
            ln = ln.strip()
            # 取含"大学/学院/学校"的校名片段
            for m in re.findall(r"[\u4e00-\u9fa5]{2,}(?:大学|学院|学校|分校)", ln):
                found.add(m)
        print(f"[A2] {label}: 提取候选校名 {len(found)}")
        return found

    sets = {
        "is_985": names_in_pdf("国家985工程大学（39）.pdf", "985"),
        "is_211": names_in_pdf("国家211工程大学（115）.pdf", "211"),
        "is_dfc": names_in_pdf("国家双一流工程大学（147）.pdf", "双一流"),
        "小985": names_in_pdf("985工程优势学科创新平台（小985）.pdf", "小985"),
        "小211": names_in_pdf("国家中西部高校基础能力建设工程（小211）.pdf", "小211"),
        "部委直属": names_in_pdf("国家部委直属院校.pdf", "部委直属"),
    }

    # 标记布尔 + 聚合 tags
    tag_of = {"is_985": "985", "is_211": "211", "is_dfc": "双一流",
              "小985": "小985", "小211": "小211", "部委直属": "部委直属"}
    matched = {k: 0 for k in sets}
    cur.execute("SELECT code, name FROM school_profiles")
    rows = cur.fetchall()
    for code, name in rows:
        tags = []
        for key, names in sets.items():
            if name in names:
                matched[key] += 1
                if key in ("is_985", "is_211", "is_dfc"):
                    cur.execute(f"UPDATE school_profiles SET {key}=TRUE WHERE code=%s", (code,))
                tags.append(tag_of[key])
        if tags:
            cur.execute("UPDATE school_profiles SET tags=%s WHERE code=%s", (";".join(tags), code))
    conn.commit()
    for k, v in matched.items():
        print(f"[A2] {k}: 命中 {v} 所")


# ---------- A3: 保研367所 ----------
def load_a3(conn):
    path = os.path.join(BASE, "全国具有保研资格的高校名单（367所）.pdf")
    lines = _pdf_lines(path, max_pages=20)
    names = set()
    for ln in lines:
        for m in re.findall(r"[\u4e00-\u9fa5]{2,}(?:大学|学院|学校|分校)", ln):
            names.add(m)
    print(f"[A3] 提取候选校名 {len(names)}")
    cur = conn.cursor()
    cur.execute("SELECT code, name FROM schools")
    name2code = {n: c for c, n in cur.fetchall()}
    hit = 0
    for n in names:
        if n in name2code:
            cur.execute("UPDATE school_profiles SET has_postgrad_recommend=TRUE WHERE code=%s", (name2code[n],))
            hit += 1
    conn.commit()
    print(f"[A3] 标记保研资格 {hit} 所")


# ---------- B6: 辽宁选考科目 xlsx ----------
def load_b6(conn):
    path = os.path.join(BASE, "2027lnzsxkap0407jx.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cur = conn.cursor()
    cur.execute("SELECT code, name FROM schools")
    name2code = {n: c for c, n in cur.fetchall()}
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    n = 0
    for row in rows[1:]:
        if not row or not row[1]:
            continue
        sch_name = str(row[1]).strip()
        major_code = str(row[2]).strip() if row[2] else None
        major_name = str(row[3]).strip() if row[3] else None
        subject_req = str(row[4]).strip() if row[4] else None
        code = name2code.get(sch_name)
        if not code or not major_name:
            continue
        # upsert 到 major_profiles（按 院校代码+专业名+年份 维度）
        cur.execute("""
            INSERT INTO major_profiles (school_code, major_code, major_name, year, subject_req, source, enriched_at)
            VALUES (%s,%s,%s,2025,%s,'ln_zk_xlsx',now())
            ON CONFLICT (school_code, major_code, major_name, year, category)
            DO UPDATE SET subject_req=EXCLUDED.subject_req, source='ln_zk_xlsx', enriched_at=now()
        """, (code, major_code, major_name, subject_req))
        n += 1
    conn.commit()
    print(f"[B6] 写入/更新 major_profiles.subject_req {n} 条")


# ---------- B7: 专业目录 2026（纯文本层级格式）----------
def load_b7(conn):
    path = os.path.join(BASE, "普通高等学校本科专业目录（2026年）.pdf")
    cur = conn.cursor()
    n = 0
    discipline = None
    category = None
    with pdfplumber.open(path) as pdf:
        for pi in range(len(pdf.pages)):
            txt = pdf.pages[pi].extract_text() or ""
            for line in txt.splitlines():
                line = line.strip()
                m_d = re.match(r"^(\d{2})\s+学科门类：(.+)$", line)
                if m_d:
                    discipline = m_d.group(2).strip()
                    category = None
                    continue
                m_c = re.match(r"^(\d{4})\s+(.+类)$", line)
                if m_c and not re.match(r"^\d{6}", line):
                    category = m_c.group(2).strip()
                    continue
                m_m = re.match(r"^(\d{6}[A-Z]?)\s+(.+)$", line)
                if m_m:
                    code = m_m.group(1).strip()
                    name = re.sub(r"\s*（注：.*?）", "", m_m.group(2)).strip()
                    if not name or name.endswith("类"):
                        continue
                    cur.execute("""
                        INSERT INTO major_catalog (code, name, category, discipline, year)
                        VALUES (%s,%s,%s,%s,2026)
                        ON CONFLICT (code, year) DO UPDATE SET name=EXCLUDED.name, category=EXCLUDED.category, discipline=EXCLUDED.discipline
                    """, (code, name, category, discipline))
                    n += 1
    conn.commit()
    print(f"[B7] 写入 major_catalog {n} 条")


def report(conn):
    cur = conn.cursor()
    print("\n=== 覆盖率报告 ===")
    cur.execute("SELECT count(*) FROM school_profiles")
    total = cur.fetchone()[0]
    for col in ["school_id_code", "affiliation", "province", "city", "level", "tags"]:
        cur.execute(f"SELECT count(*) FROM school_profiles WHERE {col} IS NOT NULL")
        c = cur.fetchone()[0]
        print(f"  school.{col:22s}: {c}/{total} ({100*c//total if total else 0}%)")
    cur.execute("SELECT count(*) FROM school_profiles WHERE has_postgrad_recommend=TRUE")
    c = cur.fetchone()[0]
    print(f"  school.has_postgrad=TRUE : {c}/{total} ({100*c//total if total else 0}%)")
    for col in ["is_985", "is_211", "is_dfc"]:
        cur.execute(f"SELECT count(*) FROM school_profiles WHERE {col}=TRUE")
        c = cur.fetchone()[0]
        print(f"  school.{col}=TRUE      : {c}")
    cur.execute("SELECT count(*) FROM major_profiles")
    mt = cur.fetchone()[0]
    cur.execute("SELECT count(*) FROM major_profiles WHERE subject_req IS NOT NULL")
    ms = cur.fetchone()[0]
    print(f"  major.subject_req        : {ms}/{mt} ({100*ms//mt if mt else 0}%)")
    cur.execute("SELECT count(*) FROM major_catalog")
    print(f"  major_catalog 总数       : {cur.fetchone()[0]}")


def main():
    ap = argparse.ArgumentParser()
    for k in ["a1", "a2", "a3", "b6", "b7", "all", "report"]:
        ap.add_argument(f"--{k}", action="store_true")
    args = ap.parse_args()
    conn = conn_db()
    if args.all or args.a1: load_a1(conn)
    if args.all or args.a2: load_a2(conn)
    if args.all or args.a3: load_a3(conn)
    if args.all or args.b6: load_b6(conn)
    if args.all or args.b7: load_b7(conn)
    if args.report or args.all: report(conn)
    conn.close()


if __name__ == "__main__":
    main()

import os, re, glob, subprocess, tempfile
import xml.etree.ElementTree as ET
import zipfile
import openpyxl
import xlrd
from config import BASE_DIR, DATA_DIRS, OCR_LANGS, OCR_DPI


class EncryptedFile(Exception):
    pass


def _xlsx_salvage(path):
    """openpyxl 打不开时的兜底：直接用 zipfile+XML 解析 xlsx。

    适用于 zip 头带加密标志位(flag 0x2)但内容未真加密、标准库 zipfile
    可完整读取、openpyxl 却报 BadZipFile 的官方导出文件
    （2024 提前批/专科批艺术类 4 个文件即此类）。
    返回 [(sheet_name, rows)]；读不了才抛 EncryptedFile。"""
    NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    NS_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    try:
        zf = zipfile.ZipFile(path)
    except Exception as e:
        raise EncryptedFile(f"{path}: zip 无法打开 {e}")
    with zf:
        names = set(zf.namelist())
        if "xl/workbook.xml" not in names:
            raise EncryptedFile(f"{path}: 非标准 xlsx 结构")
        # 共享字符串表
        sst = []
        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(f"{NS_MAIN}si"):
                sst.append("".join(t.text or "" for t in si.iter(f"{NS_MAIN}t")))
        # 工作表名：r:id -> name
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_map = {r.get("Id"): r.get("Target").lstrip("/")
                   for r in rels}
        sheets = []
        for s in wb.iter(f"{NS_MAIN}sheet"):
            tgt = rel_map.get(s.get(f"{NS_R}id"), "")
            if not tgt.startswith("xl/"):
                tgt = "xl/" + tgt
            sheets.append((s.get("name"), tgt))

        def cell_value(c):
            t = c.get("t")
            if t == "s":
                v = c.find(f"{NS_MAIN}v")
                return sst[int(v.text)] if v is not None and v.text else None
            if t == "inlineStr":
                is_el = c.find(f"{NS_MAIN}is")
                if is_el is not None:
                    return "".join(tt.text or "" for tt in is_el.iter(f"{NS_MAIN}t"))
                return None
            v = c.find(f"{NS_MAIN}v")
            if v is None or v.text is None:
                return None
            txt = v.text.strip()
            try:
                return float(txt) if ("." in txt or "e" in txt.lower()) else int(txt)
            except ValueError:
                return txt

        def col_idx(ref):
            n = 0
            for ch in ref:
                if ch.isalpha():
                    n = n * 26 + (ord(ch.upper()) - 64)
                else:
                    break
            return n - 1

        out = []
        for sheet_name, target in sheets:
            if target not in names:
                continue
            root = ET.fromstring(zf.read(target))
            rows = []
            for row in root.iter(f"{NS_MAIN}row"):
                cells = {}
                for c in row.findall(f"{NS_MAIN}c"):
                    ref = c.get("r") or ""
                    if ref:
                        cells[col_idx(ref)] = cell_value(c)
                if not cells:
                    rows.append([])
                    continue
                width = max(cells) + 1
                rows.append([cells.get(i) for i in range(width)])
            out.append((sheet_name, rows))
        if not out:
            raise EncryptedFile(f"{path}: 兜底解析未得到任何工作表")
        return out


def iter_files():
    for d in DATA_DIRS:
        root = os.path.join(BASE_DIR, d)
        for p in sorted(glob.glob(os.path.join(root, "*"))):
            if p.endswith(":Zone.Identifier"):
                continue
            base = os.path.basename(p)
            # 跳过 Excel 临时锁文件（~$ 前缀）等非数据文件
            if base.startswith("~$"):
                continue
            ext = os.path.splitext(p)[1].lower()
            if ext in (".xlsx", ".xls", ".pdf"):
                yield p


def _ole2_decrypt(path):
    """OLE2 容器 + EncryptedPackage 的官方默认密钥加密文件，解密后返回字节流。

    辽宁招考办部分官方导出文件用 Excel 默认密钥 VelvetSweatshop 加密：
    用户打开时无需输密码（Excel 静默解密），但 openpyxl/xlrd 直接打不开。
    真实用户密码加密的文件会解密失败，抛 EncryptedFile。"""
    import io
    try:
        import msoffcrypto
    except ImportError as e:
        raise EncryptedFile(f"{path}: OLE2 加密文件需 msoffcrypto-tool ({e})")
    try:
        with open(path, "rb") as f:
            of = msoffcrypto.OfficeFile(f)
            if not of.is_encrypted():
                raise EncryptedFile(f"{path}: OLE2 容器但无加密流，格式异常")
            of.load_key(password="VelvetSweatshop")
            buf = io.BytesIO()
            of.decrypt(buf)
        buf.seek(0)
        return buf
    except EncryptedFile:
        raise
    except Exception as e:
        raise EncryptedFile(f"{path}: 默认密钥解密失败（可能有真实密码）{e}")


def read_spreadsheet(path):
    """返回 [(sheet_name, rows)]；加密/损坏文件抛 EncryptedFile。"""
    ext = os.path.splitext(path)[1].lower()
    out = []
    if ext == ".xlsx":
        with open(path, "rb") as f:
            magic = f.read(4)
        if magic == b"\xd0\xcf\x11\xe0":
            # OLE2 容器：官方默认密钥加密文件，解密后按普通 xlsx 读
            wb = openpyxl.load_workbook(_ole2_decrypt(path),
                                        read_only=True, data_only=True)
        else:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            except Exception:
                # 兜底：zip 头带加密标志位但内容未加密的官方导出文件，
                # openpyxl 报 BadZipFile，改用 zipfile+XML 直读。
                return _xlsx_salvage(path)
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            out.append((ws.title, rows))
        wb.close()
    elif ext == ".xls":
        wb = xlrd.open_workbook(path)
        for sh in wb.sheets():
            rows = [[sh.cell(r, c).value for c in range(sh.ncols)]
                    for r in range(sh.nrows)]
            out.append((sh.name, rows))
    return out


def _ocr_png(png):
    r = subprocess.run(["tesseract", png, "stdout", "-l", OCR_LANGS],
                       capture_output=True, text=True)
    return r.stdout


def _pdf_textlayer_pages(path):
    """PDF 文字层提取（过滤大字竖排水印），不可用返回 None。

    辽宁招考办 PDF 带竖排单字水印（字号约 76pt，正文 10-13pt），
    直接 extract_text 会把水印字插进各行；按字号过滤后即为干净原文。
    """
    try:
        import pdfplumber
    except ImportError:
        return None
    try:
        out = []
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                filtered = page.filter(
                    lambda o: o["object_type"] != "char" or o["size"] <= 30)
                out.append((i, filtered.extract_text() or ""))
        return out
    except Exception:
        return None


def _ocr_pages(path):
    """渲染每页为 PNG 并 OCR，返回 [(page_no, text)]。"""
    pages = []
    with tempfile.TemporaryDirectory() as td:
        base = os.path.join(td, "p")
        subprocess.run(["pdftoppm", "-r", str(OCR_DPI), "-png", path, base],
                       check=True)
        for png in sorted(glob.glob(base + "-*.png")):
            num = int(re.search(r"-(\d+)\.png$", png).group(1))
            pages.append((num, _ocr_png(png)))
    return pages


def read_pdf(path):
    """文字层优先，逐页不足（扫描/文字层过少）时回退 OCR。"""
    tl = _pdf_textlayer_pages(path)
    if tl:
        pages = []
        need_ocr = False
        for num, text in tl:
            if len(text.strip()) < 300:   # 文字层过少，视为扫描页
                need_ocr = True
                break
            pages.append((num, text))
        if not need_ocr:
            return pages
    return _ocr_pages(path)

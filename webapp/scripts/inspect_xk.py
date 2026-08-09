# -*- coding: utf-8 -*-
"""临时检查：2027 辽宁选考科目要求三表（bk/zk/jx）结构抽样"""
import openpyxl, sys

BASE = "/home/ekewang/projects/gaokao/ln/2026allmaterial/"
for tag in ("bk", "zk", "jx"):
    path = BASE + f"2027lnzsxkap0407{tag}.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"===== {tag} sheets={wb.sheetnames}")
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        head = [next(rows, None) for _ in range(4)]
        n = ws.max_row
        print(f"-- sheet [{ws.title}] rows={n} cols={ws.max_column}")
        for h in head:
            print("   ", [str(c)[:28] if c is not None else None for c in h])
        # 抽样中部两行
        import itertools
        for r in itertools.islice(rows, 6, 8):
            print("   ...", [str(c)[:28] if c is not None else None for c in r])
    wb.close()

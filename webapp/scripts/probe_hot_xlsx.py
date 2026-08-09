# -*- coding: utf-8 -*-
"""查 热门大学.xlsx 中是否出现 非酒亦酒/非九亦九 等词"""
import openpyxl, glob, os

path = "/home/ekewang/projects/gaokao/ln/2026allmaterial/热门大学介绍/热门大学.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
hits = []
for ws in wb.worksheets:
    print(f"== sheet: {ws.title} ({ws.max_row}x{ws.max_column})")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        for cell in row:
            if isinstance(cell, str) and ("酒" in cell or "非九" in cell or "亦九" in cell):
                hits.append((ws.title, i, cell[:120]))
        if i <= 3:
            print("  行样例:", [str(c)[:20] if c is not None else "" for c in row][:8])
print("\n命中:", len(hits))
for h in hits[:30]:
    print(h)

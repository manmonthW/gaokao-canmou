"""志愿方案 xlsx 导出（Phase 3）。

列结构对齐辽宁「专业+学校」平行志愿（roadmap §三）：
  序号 / 档位 / 院校代码 / 院校名称 / 专业代码 / 专业名称 /
  往年最低分 / 往年最低位次 / 位次差 / 层次 / 城市

说明：
  - 方案条目由前端传入（含创建时的数据版本与风险快照），后端只负责排版生成文件，
    不重新计算风险 —— 保证导出内容与用户看到的快照一致。
  - 导出默认不包含用户真实姓名（spec §隐私）。
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "序号", "档位", "院校代码", "院校名称", "专业代码", "专业名称",
    "往年最低分", "往年最低位次", "位次差", "层次", "城市", "备注",
]
COL_WIDTHS = [6, 8, 10, 26, 10, 30, 11, 12, 14, 10, 12, 24]

RISK_FILL = {
    "冲": "FDE9D9",   # 橙
    "稳": "DCE6F1",   # 蓝
    "保": "EBF1DE",   # 绿
    "高波动": "F2DCDB",  # 红
    "数据不足": "F2F2F2",  # 灰
}

_thin = Side(style="thin", color="D0D0D0")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _diff_text(d):
    if d is None:
        return "—"
    if d < 0:
        return f"领先 {abs(d)}"
    if d > 0:
        return f"落后 {d}"
    return "持平"


def build_plan_xlsx(payload: dict) -> bytes:
    """payload:
    {
      plan_name, note, data_version, created_at,
      examinee: {year, category, subject, batch, score, rank},
      items: [{risk, school_code, school_name, major_code, major_name,
               last_year, last_year_score, last_year_rank, rank_diff_last,
               level, city, note}, ...]
    }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "志愿草案"

    ex = payload.get("examinee") or {}
    plan_name = payload.get("plan_name") or "志愿方案"

    # ---- 抬头区 ----
    ncols = len(HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"{plan_name}（志愿草案 · 辽宁「专业+学校」平行志愿）")
    c.font = Font(size=14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    info = (
        f"考生条件：{ex.get('year', '')} 年 {ex.get('category', '')} "
        f"{ex.get('subject', '')} {ex.get('batch', '')}；"
        f"分数 {ex.get('score') if ex.get('score') is not None else '—'}，"
        f"位次 {ex.get('rank') if ex.get('rank') is not None else '—'}"
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=info)
    c.font = Font(size=10, color="555555")

    meta = (
        f"数据版本：{payload.get('data_version') or '—'}；"
        f"方案创建：{payload.get('created_at') or '—'}；"
        f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(row=3, column=1, value=meta)
    c.font = Font(size=10, color="555555")

    # ---- 表头 ----
    head_row = 5
    for j, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=head_row, column=j, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border
    ws.row_dimensions[head_row].height = 20

    # ---- 条目 ----
    items = payload.get("items") or []
    r = head_row + 1
    for i, it in enumerate(items, start=1):
        risk = it.get("risk") or ""
        vals = [
            i,
            risk,
            it.get("school_code"),
            it.get("school_name"),
            it.get("major_code") or "—",
            it.get("major_name"),
            it.get("last_year_score") if it.get("last_year_score") is not None else "—",
            it.get("last_year_rank") if it.get("last_year_rank") is not None else "—",
            _diff_text(it.get("rank_diff_last")),
            it.get("level") or "—",
            it.get("city") or "—",
            it.get("note") or "",
        ]
        fill = RISK_FILL.get(risk)
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(size=10)
            cell.border = _border
            if j in (1, 2, 3, 5, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill and j == 2:
                cell.fill = PatternFill("solid", fgColor=fill)
        r += 1

    # ---- 免责声明 ----
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(
        row=r, column=1,
        value="说明：往年最低分/位次为该单元最近一年投档数据；「档位」为规则模型初步判定（不含概率）。"
              "本表仅供参考，最终请以辽宁省招考部门及院校官方发布为准。",
    )
    c.font = Font(size=9, color="888888")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28

    for j, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{head_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
